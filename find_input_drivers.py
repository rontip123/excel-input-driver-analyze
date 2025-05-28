import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter
import re
from collections import defaultdict
import xlrd
import os
import datetime

def get_cell_value(cell):
    """Safely get cell value, handling None and other types"""
    if cell.value is None:
        return ""
    return str(cell.value)

def is_text_only(value):
    """
    Check if a cell contains only text or dates (no formulas).
    Returns True if the value is a string, date, or datetime and doesn't contain formula operators.
    Allows:
    - hyphens between letters (e.g., "co-branded")
    - hyphens with spaces on each side (e.g., "A - B")
    - dates and datetimes
    but treats standalone minus signs as operators.
    """
    if value is None:
        return False
    
    # Allow dates and datetimes
    if isinstance(value, (datetime.datetime, datetime.date)):
        return True
    
    # If it's already a string, check if it contains formula operators
    if isinstance(value, str):
        # First, replace any hyphens between letters or with spaces with a placeholder
        value_with_placeholder = re.sub(r'(?<=[a-zA-Z])-(?=[a-zA-Z])|(?<=\s)-(?=\s)', 'HYPHEN_PLACEHOLDER', value)
        
        # Only exclude true formula operators, excluding "/" as it can be used for "or"
        formula_operators = ['=', '+', '-', '*', '!']
        return not any(op in value_with_placeholder for op in formula_operators)
    
    # For other types (numbers, etc.), return False
    return False

def is_number_only_formula(value):
    """Check if a formula contains only numbers and basic arithmetic operators"""
    if not value or not isinstance(value, str) or not value.startswith('='):
        return False
    # Remove the equals sign and whitespace
    formula = value[1:].strip()
    # Remove all numbers, decimal points, and basic arithmetic operators
    cleaned = re.sub(r'[\d\.\+\-\*\/\(\)]', '', formula)
    # If nothing remains, it's a number-only formula
    return not cleaned

def is_input_driver_formula(value):
    """Check if a formula is a valid input driver (contains numbers and/or cell references)"""
    if not value or not isinstance(value, str) or not value.startswith('='):
        return False
    
    # Remove the equals sign and whitespace
    formula = value[1:].strip()
    
    # List of common Excel functions that can be used in input drivers
    common_functions = ['SUM', 'AVERAGE', 'MAX', 'MIN', 'COUNT']
    
    # Remove all numbers, decimal points, basic arithmetic operators, and cell references
    cleaned = re.sub(r'[\d\.\+\-\*\/\(\)]', '', formula)
    cleaned = re.sub(r'[A-Za-z]+[0-9]+', '', cleaned)
    
    # Remove common Excel functions
    for func in common_functions:
        cleaned = cleaned.replace(func, '')
    
    # If nothing remains or only commas and colons remain (used in ranges), it's an input driver formula
    cleaned = cleaned.replace(',', '').replace(':', '')
    return not cleaned

def find_row_label(sheet, row, col):
    """
    Find the row label by searching from the column immediately to the left of the input driver cell,
    moving left until we find a text cell or reach column A.
    Returns the first text-only cell found, or blank if none found.
    """
    # Start from the column immediately to the left of the input driver cell
    start_col = col - 1
    
    # Search from right to left until we reach column A
    for current_col in range(start_col, 0, -1):  # 0 is column A
        cell = sheet.cell(row=row, column=current_col)
        value = cell.value
        
        if is_text_only(value):
            return value
    
    return ""  # Return blank if no text cell found

def find_column_label(sheet, row, col):
    """Find the first text-only label by scanning up"""
    start_row = row - 1
    # Start from the given row and move up
    for row in range(start_row, 0, -1):
        cell = sheet.cell(row=row, column=col)
        if cell.value and is_text_only(cell.value):
            return get_cell_value(cell)
    return ""

def get_cell_reference(row, col):
    """Convert row and column numbers to Excel cell reference"""
    return f"{get_column_letter(col)}{row}"

def get_cell_value_xlrd(cell, workbook):
    """Safely get cell value from xlrd, handling different types"""
    if cell.ctype == xlrd.XL_CELL_EMPTY:
        return ""
    elif cell.ctype == xlrd.XL_CELL_NUMBER:
        # Check if it's a date
        if xlrd.datemode.XLDate(cell.value, workbook.datemode) is not None:
            return str(cell.value)
        return str(cell.value)
    elif cell.ctype == xlrd.XL_CELL_TEXT:
        return cell.value
    elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
        return str(cell.value)
    elif cell.ctype == xlrd.XL_CELL_ERROR:
        return ""
    elif cell.ctype == xlrd.XL_CELL_BLANK:
        return ""
    return str(cell.value)

def is_text_only_xlrd(cell, workbook):
    """Check if a cell contains only text using xlrd"""
    if cell.ctype == xlrd.XL_CELL_EMPTY or cell.ctype == xlrd.XL_CELL_BLANK:
        return False
    if cell.ctype == xlrd.XL_CELL_TEXT:
        value = cell.value
        return not any(op in str(value) for op in ['+', '-', '*', '/', '(', ')', '!', '='])
    return False

def find_row_label_xlrd(sheet, row, start_col, workbook):
    """Find the first text-only label by scanning left using xlrd"""
    try:
        print(f"\nSearching for row label at row {row}, starting from column {start_col}:")
        # Start from the given column and move left
        for col in range(start_col, -1, -1):
            cell = sheet.cell(row, col)
            cell_value = get_cell_value_xlrd(cell, workbook)
            print(f"  Checking column {get_column_letter(col+1)}: value='{cell_value}', is_text={is_text_only_xlrd(cell, workbook)}")
            if cell.ctype != xlrd.XL_CELL_EMPTY and is_text_only_xlrd(cell, workbook):
                print(f"  Found row label: '{cell_value}' at column {get_column_letter(col+1)}")
                return cell_value
        print("  No row label found after scanning all columns")
        return ""
    except Exception as e:
        print(f"  Error while searching for row label: {str(e)}")
        return ""

def find_column_label_xlrd(sheet, row, col, workbook):
    """Find the first text-only label by scanning up using xlrd"""
    start_row = row - 1
    for row in range(start_row, -1, -1):
        cell = sheet.cell(row, col)
        if cell.ctype != xlrd.XL_CELL_EMPTY and is_text_only_xlrd(cell, workbook):
            return get_cell_value_xlrd(cell, workbook)
    return ""

def find_input_drivers(excel_file, output_csv):
    try:
        file_ext = os.path.splitext(excel_file)[1].lower()
        
        if file_ext == '.xls':
            # Use xlrd for .xls files
            workbook = xlrd.open_workbook(excel_file)
            sheet = workbook.sheet_by_name('FY Financials')
            
            # List of cells to check
            cells_to_check = ['CP176', 'CP185', 'CP187', 'CP204', 'CP213', 'CP234', 'CP235', 'CP243', 'CP244']
            
            print("\nChecking specific cells in FY Financials:")
            for cell_ref in cells_to_check:
                try:
                    # Convert Excel-style reference to row/col
                    col = ''.join(filter(str.isalpha, cell_ref))
                    row = int(''.join(filter(str.isdigit, cell_ref))) - 1  # xlrd is 0-based
                    col_idx = openpyxl.utils.column_index_from_string(col) - 1  # xlrd is 0-based
                    
                    cell = sheet.cell(row, col_idx)
                    print(f"\n{'='*50}")
                    print(f"Analyzing cell {cell_ref}:")
                    print(f"Value: {get_cell_value_xlrd(cell, workbook)}")
                    print(f"Type: {cell.ctype}")
                    
                    row_label = find_row_label_xlrd(sheet, row, col_idx - 1, workbook)
                    column_label = find_column_label_xlrd(sheet, row - 1, col_idx, workbook)
                    
                    print(f"Row label: {row_label}")
                    print(f"Column label: {column_label}")
                    
                    # For xlrd, we'll consider a cell an input driver if it's a number or text
                    is_input = cell.ctype in [xlrd.XL_CELL_NUMBER, xlrd.XL_CELL_TEXT]
                    print(f"Would be input driver: {is_input}")
                    
                except Exception as e:
                    print(f"Error checking {cell_ref}: {str(e)}")
            
        else:
            # Use openpyxl for .xlsx files
            wb = openpyxl.load_workbook(excel_file, data_only=False)
            
            # Dictionary to store ALL input drivers by sheet
            all_input_drivers_by_sheet = defaultdict(list)
            
            # Regular expression to find cell references in formulas
            cell_ref_pattern = r'[A-Za-z]+[0-9]+'
            
            # Iterate through each sheet
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                
                # Step 1: Find all cells that contain formulas
                formula_cells = []
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith('='):
                            formula_cells.append(cell)
                
                # Step 2: Find all cells that are referenced by formulas
                referenced_cells = set()
                for cell in formula_cells:
                    if cell.value:
                        # Find all cell references in the formula
                        refs = re.findall(cell_ref_pattern, cell.value)
                        referenced_cells.update(refs)
                
                # Step 3: Check each referenced cell to see if it's an input driver
                for cell_ref in referenced_cells:
                    try:
                        # Convert cell reference to coordinates
                        col = ''.join(filter(str.isalpha, cell_ref))
                        row = int(''.join(filter(str.isdigit, cell_ref)))
                        col_idx = openpyxl.utils.column_index_from_string(col)
                        
                        # Get the cell value
                        cell = sheet[f"{col}{row}"]
                        
                        # Check if this referenced cell is an input driver
                        is_input = cell.value is not None and (
                            not isinstance(cell.value, str) or
                            not cell.value.startswith('=') or
                            is_number_only_formula(cell.value) or
                            is_input_driver_formula(cell.value)
                        )
                        
                        if is_input:
                            # Find row label by scanning left
                            row_label = find_row_label(sheet, row, col_idx - 1)
                            
                            # Only process if row_label is not empty
                            if row_label:
                                # Find column label by scanning up
                                column_label = find_column_label(sheet, row - 1, col_idx)
                                
                                # Only process if column_label is not empty
                                if column_label:
                                    # Store input driver information
                                    all_input_drivers_by_sheet[sheet_name].append(
                                        (row_label, col_idx, row, cell_ref, str(cell.value), column_label)
                                    )
                    except Exception as e:
                        continue
            
            # Process each sheet to find the column that is rightmost for the most row_labels
            final_input_drivers = []
            
            for sheet_name, input_drivers in all_input_drivers_by_sheet.items():
                sheet = wb[sheet_name]
                
                # Find the selected column for this sheet
                row_label_rightmost_cols = defaultdict(int)
                for row_label, col_idx, _, _, _, _ in input_drivers:
                    if col_idx > row_label_rightmost_cols[row_label]:
                        row_label_rightmost_cols[row_label] = col_idx
                
                col_counts = defaultdict(int)
                for rightmost_col in row_label_rightmost_cols.values():
                    col_counts[rightmost_col] += 1
                
                if not col_counts:
                    continue
                
                max_count = max(col_counts.values())
                max_cols = [col for col, count in col_counts.items() if count == max_count]
                selected_col = max(max_cols)
                
                # Only include input drivers from the selected column
                for row_label, col_idx, row, cell_ref, content, column_label in input_drivers:
                    if col_idx == selected_col:  # Only process cells in the selected column
                        # Get information about cells to the right
                        first_right_col = col_idx + 1
                        second_right_col = col_idx + 2
                        
                        # Get cell references for cells to the right
                        first_right_cell = get_cell_reference(row, first_right_col)
                        second_right_cell = get_cell_reference(row, second_right_col)
                        
                        # Use find_column_label function to get column labels for the right cells
                        first_right_col_label = find_column_label(sheet, row, first_right_col)
                        second_right_col_label = find_column_label(sheet, row, second_right_col)
                        
                        final_input_drivers.append({
                            'Sheet': sheet_name,
                            'Row_Label': row_label,
                            'Row': row,
                            'Cell': cell_ref,
                            'Content': content,
                            'Column_Label': column_label,
                            'FirstRight_Cell': first_right_cell,
                            'FirstRight_Column_Label': first_right_col_label,
                            'SecondRight_Cell': second_right_cell,
                            'SecondRight_Column_Label': second_right_col_label
                        })
            
            # Create DataFrames and save outputs
            if final_input_drivers:
                # Create DataFrame with all input drivers
                df = pd.DataFrame(final_input_drivers)
                df = df.sort_values(['Sheet', 'Row'])
                df = df.drop(['Row'], axis=1)
                df.to_csv(output_csv, index=False)
                print(f"Found {len(final_input_drivers)} input drivers.")
                print(f"Results saved to {output_csv}")
            else:
                print("No input drivers found in the workbook.")
            
    except Exception as e:
        print(f"An error occurred: {str(e)}")

def main():
    import sys
    
    if len(sys.argv) != 3:
        print("Usage: python find_input_drivers.py <input_excel_file> <output_csv_file>")
        sys.exit(1)
    
    input_excel = sys.argv[1]
    output_csv = sys.argv[2]
    
    find_input_drivers(input_excel, output_csv)

if __name__ == "__main__":
    main() 